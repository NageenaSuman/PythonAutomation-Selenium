""" Importing the user input for instrument name from Excel File """
import warnings
import xml.etree.ElementTree as ET
import re
import openpyxl

# Ignoring the data validation error
warnings.simplefilter(action='ignore', category=UserWarning)


class XML:
    """ Creating class to load XML config file for URL and login credentials """

    # Parsing the ScriptConfig(Userdata) XML file
    Tree = ET.parse("Files\\Config.xml")
    Root = Tree.getroot()
    url = Root[0].text  # AOL Site Link
    user_name = Root[1].text  # Username
    password = Root[2].text  # Password


class Data:
    """ Converting the raw instrument to title case """

    instru_list = []
    instru = []
    char_list = []
    Workbook = openpyxl.load_workbook("Files\\UserFile.xlsx")
    sheet = Workbook.active
    for i in range(2, sheet.max_row + 1):
        instrument_name = sheet.cell(row=i, column=2).value
        if instrument_name is not None:
            instrument = re.sub("[^A-Za-z0-9]", "", instrument_name).lower()
            FirstChar = instrument[0].upper()
            # Finalising instrument name
            instru_list.append(instrument)
            char_list.append(FirstChar)
            instru.append(instrument_name)

        else:
            instrument = None
            instru_list.append(None)
            char_list.append(None)
